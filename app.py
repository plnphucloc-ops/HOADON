import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO

st.set_page_config(layout="wide")
st.title("🚐 PHẦN MỀM TẠO FILE HÓA ĐƠN ĐIỆN TỬ")

# ================== DATA TUYẾN ==================
routes = {
    "DL-GL": {
        "07:00": "49H-046.85",
        "10:00": "49G-000.71",
        "17:00": "49B-019.00"
    },
    "GL-DL": {
        "07:00": "49H-046.85",
        "13:00": "49G-000.71",
        "17:00": "49B-019.00"
    },
    "BMT-DL": {
        "07:00": "49B-013.18"
    },
    "DL-BMT": {
        "13:00": "49B-013.18"
    }
}

# ================== GIÁ ==================
gia_tuyen = {
    "DL-GL": 420000,
    "GL-DL": 420000,
    "BMT-DL": 300000,
    "DL-BMT": 300000
}

# ================== XE ==================
all_cars = [
    "49B-016.93",
    "49B-017.39",
    "49B-019.00",
    "49G-000.71",
    "49B-013.18",
    "49H-046.85"
]

# ================== CHỌN ==================
colA, colB, colC = st.columns(3)

with colA:
    tuyen = st.selectbox("🚐 Tuyến", list(routes.keys()))

with colB:
    gio = st.selectbox("⏰ Giờ", list(routes[tuyen].keys()))

with colC:
    xe_mac_dinh = routes[tuyen][gio]
    options = ["--- Không chọn ---"] + all_cars
    index = options.index(xe_mac_dinh) if xe_mac_dinh in options else 0
    xe = st.selectbox("🚌 Số xe", options, index=index)

# ================== NGÀY ==================
ngay = st.date_input("📅 Ngày chạy")
ngay_file = ngay.strftime("%d.%m.%Y")
ngay_show = ngay.strftime("%d/%m/%Y")
gio_clean = gio.replace(":", "H")

# ================== FORM ==================
st.divider()
st.subheader("🧾 Nhập thông tin vé")

gia_1ve = gia_tuyen[tuyen]
st.info(f"💰 Giá tuyến {tuyen}: {gia_1ve:,} đ / vé")

with st.form("form_ve"):
    col1, col2 = st.columns(2)

    with col1:
        ten = st.text_area("Họ tên khách / Đơn vị", height=120)
        cccd = st.text_input("CCCD / MST")
        sdt = st.text_input("Số điện thoại")
        so_ve = st.number_input("Số vé", min_value=1, value=1)

    with col2:
        st.text_input("Giá 1 vé", value=f"{gia_1ve:,} đ", disabled=True)
        thanh_tien = so_ve * gia_1ve
        st.text_input("Thành tiền", value=f"{thanh_tien:,} đ", disabled=True)

    submit = st.form_submit_button("➕ Thêm vé")

# ================== DATA ==================
if "ds_ve" not in st.session_state:
    st.session_state.ds_ve = []

if submit:
    if xe == "--- Không chọn ---":
        st.warning("⚠️ Vui lòng chọn xe")
    else:
        st.session_state.ds_ve.append({
            "ten": ten,
            "cccd": cccd,
            "sdt": sdt,
            "gio": gio,
            "tuyen": tuyen,
            "xe": xe,
            "so_ve": so_ve,
            "gia": thanh_tien
        })

# ================== DANH SÁCH + XÓA ==================
st.divider()
st.subheader("📋 Danh sách vé")

if st.session_state.ds_ve:
    df = pd.DataFrame(st.session_state.ds_ve)

    for i, row in df.iterrows():
        col1, col2 = st.columns([10, 1])

        with col1:
            st.write(
                f"👤 {row['ten']} | 📞 {row['sdt']} | 🚐 {row['tuyen']} | ⏰ {row['gio']} | 🚌 {row['xe']} | 🎫 {row['so_ve']} | 💰 {row['gia']:,} đ"
            )

        with col2:
            if st.button("❌", key=f"xoa_{i}"):
                st.session_state.ds_ve.pop(i)
                st.rerun()

    # tổng tiền
    tong = sum([x["gia"] for x in st.session_state.ds_ve])
    st.success(f"💰 Tổng tiền: {tong:,} đ")

    # xóa tất cả
    if st.button("🗑️ Xóa tất cả"):
        st.session_state.ds_ve = []
        st.rerun()

# ================== XUẤT FILE ==================
def tao_file():
    wb = Workbook()
    ws = wb.active

    # ===== STYLE =====
    font_title = Font(name="Times New Roman", size=16, bold=True)
    font_header = Font(name="Times New Roman", size=12, bold=True)
    font_normal = Font(name="Times New Roman", size=12)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    fill = PatternFill(start_color="EEEEEE", fill_type="solid")

    # ===== TIÊU ĐỀ =====
    ws.merge_cells("A1:H1")
    ws["A1"] = "CÔNG TY PHÚC HẢI ĐÀ LẠT"
    ws["A1"].font = font_title
    ws["A1"].alignment = center

    ws.merge_cells("A2:H2")
    ws["A2"] = f"TUYẾN {tuyen} | GIỜ {gio} | XE {xe} | NGÀY {ngay_show}"
    ws["A2"].font = font_normal
    ws["A2"].alignment = center

    # ===== HEADER =====
    headers = [
        "Họ tên khách/Tên đơn vị",
        "CCCD/MST",
        "Số điện thoại",
        "Giờ xuất bến",
        "Tuyến xe",
        "Số xe",
        "Số vé",
        "Thành tiền"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = font_header
        cell.alignment = center
        cell.fill = fill
        cell.border = thin

    ws.row_dimensions[4].height = 30

    # ===== ĐỘ RỘNG CỘT =====
    widths = [40, 22, 20, 15, 15, 15, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    # ===== DATA =====
    start = 5

    for i, row in enumerate(st.session_state.ds_ve, start=start):

        ws.cell(i,1,row["ten"])
        ws.cell(i,2,row["cccd"])
        ws.cell(i,3,row["sdt"])
        ws.cell(i,4,row["gio"])
        ws.cell(i,5,row["tuyen"])
        ws.cell(i,6,row["xe"])
        ws.cell(i,7,row["so_ve"])

        money = ws.cell(i,8,row["gia"])
        money.number_format = '#,##0 "đ"'

        for col in range(1,9):
            c = ws.cell(i,col)
            c.font = font_normal
            c.border = thin

            if col == 1:
                c.alignment = left
            elif col == 8:
                c.alignment = right
            else:
                c.alignment = center

        # ===== AUTO CHIỀU CAO DÒNG =====
        text = str(row["ten"])
        lines = text.count("\n") + 1
        ws.row_dimensions[i].height = max(28, lines * 18)

    # ===== TỔNG =====
    last = len(st.session_state.ds_ve) + start

    ws.cell(last,7,"Tổng").font = font_header
    ws.cell(last,7).alignment = center

    total = ws.cell(last,8,sum([x["gia"] for x in st.session_state.ds_ve]))
    total.number_format = '#,##0 "đ"'
    total.font = font_header

    for col in range(1,9):
        ws.cell(last,col).border = thin

    # ===== CĂN GIỮA TRANG (IN ĐẸP) =====
    ws.page_setup.fitToWidth = 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ================== DOWNLOAD ==================
if st.session_state.ds_ve:
    file_name = f"{tuyen}_{gio_clean}_{ngay_file}.xlsx"

    st.download_button(
        "📥 Xuất Excel",
        data=tao_file(),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
